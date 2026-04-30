from __future__ import annotations

import hashlib
import json
import os
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse
from uuid import uuid4

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.platform.settings.base import settings

from .schemas import ListingLoadsheetPreviewRequest


LOADSHEET_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
LOADSHEET_SUBMIT_RETRY_STATUS_CODES = {408, 409, 425, 429, 500, 502, 503, 504}
TAKEALOT_APPROVED_REVIEW_STATUSES = {
    "approved",
    "accepted",
    "passed",
    "content_approved",
    "review_approved",
    "submission_approved",
}
TAKEALOT_PARTIAL_REVIEW_STATUSES = {
    "partial",
    "partially_approved",
    "approved_with_errors",
    "completed_with_errors",
    "needs_changes",
}
TAKEALOT_REJECTED_REVIEW_STATUSES = {
    "rejected",
    "declined",
    "disapproved",
    "invalid",
    "failed",
    "failure",
    "error",
}
TAKEALOT_PENDING_REVIEW_STATUSES = {
    "queued",
    "pending",
    "submitted",
    "processing",
    "in_review",
    "under_review",
    "reviewing",
}
SENSITIVE_RESPONSE_KEY_PARTS = ("api_key", "apikey", "authorization", "token", "secret", "password", "credential")
AUTHORIZATION_VALUE_RE = re.compile(
    r"\b(authorization)\s*[:=]\s*(bearer|key|token)\s+[A-Za-z0-9._~+/=-]{8,}",
    re.IGNORECASE,
)
AUTHORIZATION_HEADER_RE = re.compile(
    r"\b(authorization)\s*[:=]\s*['\"]?[A-Za-z0-9._~+/=-]{8,}",
    re.IGNORECASE,
)
BEARER_TOKEN_RE = re.compile(r"\bBearer\s+[A-Za-z0-9._~+/=-]{8,}", re.IGNORECASE)
KEY_TOKEN_RE = re.compile(r"\b(Key|Token)\s+[A-Za-z0-9._~+/=-]{8,}", re.IGNORECASE)
SECRET_ASSIGNMENT_RE = re.compile(
    r"\b(api[_-]?key|apikey|access[_-]?token|token|secret|password|credential)\s*[:=]\s*['\"]?([^\s'\"&,;]{4,})",
    re.IGNORECASE,
)
SECRET_QUERY_RE = re.compile(
    r"([?&](?:api[_-]?key|apikey|access[_-]?token|token|secret|password|credential)=)([^&\s]+)",
    re.IGNORECASE,
)
LONG_SECRET_RE = re.compile(
    r"(?<![A-Za-z0-9._~+/=-])(?=[A-Za-z0-9._~+/=-]{40,})(?=[A-Za-z0-9._~+/=-]*[A-Za-z])(?=[A-Za-z0-9._~+/=-]*\d)[A-Za-z0-9._~+/=-]{40,}(?![A-Za-z0-9._~+/=-])"
)


class ListingLoadsheetSubmitError(RuntimeError):
    def __init__(
        self,
        message: str,
        *,
        status_code: int | None = None,
        official_response: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.official_response = official_response or {}


class ListingLoadsheetStatusError(RuntimeError):
    def __init__(
        self,
        message: str,
        *,
        status_code: int | None = None,
        official_response: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.message = message
        self.status_code = status_code
        self.official_response = official_response or {}


class ListingLoadsheetService:
    def __init__(
        self,
        *,
        storage_root: Path | None = None,
        public_base_url: str | None = None,
    ) -> None:
        repo_root = Path(__file__).resolve().parents[5]
        configured_storage_root = (
            str(storage_root)
            if storage_root is not None
            else os.getenv("XH_LISTING_FILE_STORAGE_DIR", settings.listing_file_storage_dir)
        )
        resolved_storage_root = Path(configured_storage_root)
        if not resolved_storage_root.is_absolute():
            resolved_storage_root = repo_root / resolved_storage_root
        self.repo_root = repo_root
        self.storage_root = resolved_storage_root
        self.public_base_url = (
            public_base_url
            if public_base_url is not None
            else os.getenv("XH_LISTING_LOADSHEET_PUBLIC_BASE_URL", "")
        ).strip()
        self.submit_base_url = os.getenv("XH_TAKEALOT_LOADSHEET_BASE_URL", settings.takealot_catalog_base_url).rstrip("/")
        self.submit_timeout_seconds = max(5.0, float(os.getenv("XH_TAKEALOT_LOADSHEET_TIMEOUT_SECONDS", "90")))

    def build_preview(
        self,
        *,
        request: ListingLoadsheetPreviewRequest,
        category: dict[str, Any],
        allowed_attributes: list[dict[str, Any]],
        assets: list[dict[str, Any]],
        brand: dict[str, Any] | None,
        brand_catalog_ready: bool,
    ) -> dict[str, Any]:
        validation = self.validate_payload(
            request=request,
            category=category,
            allowed_attributes=allowed_attributes,
            assets=assets,
            brand=brand,
            brand_catalog_ready=brand_catalog_ready,
        )
        asset = self._write_workbook(
            request=request,
            category=category,
            validation=validation,
        )
        return {
            "valid": validation["valid"],
            "issues": validation["issues"],
            "loadsheet_asset": asset,
            "generated_fields": validation["generated_fields"],
            "missing_required_fields": validation["missing_required_fields"],
            "warnings": validation["warnings"],
        }

    def validate_payload(
        self,
        *,
        request: ListingLoadsheetPreviewRequest,
        category: dict[str, Any],
        allowed_attributes: list[dict[str, Any]],
        assets: list[dict[str, Any]],
        brand: dict[str, Any] | None,
        brand_catalog_ready: bool,
    ) -> dict[str, Any]:
        issues: list[dict[str, str]] = []
        warnings: list[str] = []
        missing_required_fields: list[str] = []

        self._require_text(request.sku, "sku", "SKU is required.", issues, missing_required_fields)
        self._require_text(request.barcode, "barcode", "Barcode is required.", issues, missing_required_fields)
        self._require_text(request.title, "title", "Title is required.", issues, missing_required_fields)
        self._require_text(
            request.whats_in_the_box,
            "whats_in_the_box",
            "What's in the Box is required.",
            issues,
            missing_required_fields,
        )
        if request.whats_in_the_box.strip() and not self._valid_whats_in_the_box(request.whats_in_the_box):
            issues.append(
                self._issue(
                    "error",
                    "whats_in_the_box",
                    "Each line must use the exact format '<quantity> x <Product Name>'.",
                )
            )

        selling_price = self._positive_number(request.selling_price, "selling_price", issues, missing_required_fields)
        rrp = self._required_nonnegative_number(request.rrp, "rrp", issues, missing_required_fields)
        if selling_price is not None and rrp is not None and rrp < selling_price:
            issues.append(self._issue("error", "rrp", "RRP must be greater than or equal to selling_price."))
        self._required_nonnegative_int(request.stock_quantity, "stock_quantity", issues, missing_required_fields)
        self._required_nonnegative_int(
            request.minimum_leadtime_days,
            "minimum_leadtime_days",
            issues,
            missing_required_fields,
        )
        self._positive_number(request.length_cm, "length_cm", issues, missing_required_fields)
        self._positive_number(request.width_cm, "width_cm", issues, missing_required_fields)
        self._positive_number(request.height_cm, "height_cm", issues, missing_required_fields)
        self._positive_number(request.weight_g, "weight_g", issues, missing_required_fields)

        brand_id = (request.brand_id or "").strip()
        brand_name = request.brand_name.strip()
        if brand_id or brand_name:
            if not brand_catalog_ready:
                issues.append(
                    self._issue(
                        "error",
                        "brand",
                        "Takealot brand library is empty; import the brand catalog before validating brand fields.",
                    )
                )
            elif brand is None:
                issues.append(
                    self._issue(
                        "error",
                        "brand",
                        "brand_id or brand_name was supplied but did not match takealot_brands.",
                    )
                )
        else:
            warnings.append("No Takealot brand supplied; loadsheet will be generated without a brand value.")

        image_result = self._collect_public_image_urls(
            request=request,
            category=category,
            assets=assets,
            issues=issues,
        )
        if request.image_urls:
            warnings.append("Remote image accessibility is not checked during loadsheet preview.")

        dynamic_result = self._validate_dynamic_attributes(
            raw_values=request.dynamic_attributes,
            allowed_attributes=allowed_attributes,
            issues=issues,
            missing_required_fields=missing_required_fields,
        )

        generated_fields = {
            "store_id": request.store_id,
            "category_id": int(category.get("category_id") or request.category_id),
            "category_path_en": category.get("path_en") or "",
            "category_path_zh": category.get("path_zh") or "",
            "loadsheet_template_id": category.get("loadsheet_template_id"),
            "loadsheet_template_name": category.get("loadsheet_template_name") or "",
            "brand_id": (brand or {}).get("brand_id") or brand_id or None,
            "brand_name": (brand or {}).get("brand_name") or brand_name,
            "sku": request.sku.strip(),
            "barcode": request.barcode.strip(),
            "title": request.title.strip(),
            "subtitle": request.subtitle.strip(),
            "description": request.description.strip(),
            "whats_in_the_box": self._box_text(request.whats_in_the_box),
            "selling_price": selling_price,
            "rrp": rrp,
            "stock_quantity": request.stock_quantity,
            "minimum_leadtime_days": request.minimum_leadtime_days,
            "seller_warehouse_id": request.seller_warehouse_id.strip(),
            "length_cm": request.length_cm,
            "width_cm": request.width_cm,
            "height_cm": request.height_cm,
            "weight_g": request.weight_g,
            "image_urls": image_result["image_urls"],
            "dynamic_attributes": dynamic_result,
        }
        return {
            "valid": not any(issue["level"] == "error" for issue in issues),
            "issues": issues,
            "generated_fields": generated_fields,
            "missing_required_fields": self._dedupe(missing_required_fields),
            "warnings": self._dedupe(warnings),
        }

    def _write_workbook(
        self,
        *,
        request: ListingLoadsheetPreviewRequest,
        category: dict[str, Any],
        validation: dict[str, Any],
    ) -> dict[str, Any]:
        workbook = Workbook()
        product_sheet = workbook.active
        product_sheet.title = "Loadsheet"

        generated_fields = validation["generated_fields"]
        rows = [
            ("Store ID", generated_fields.get("store_id")),
            ("Category ID", generated_fields.get("category_id")),
            ("Category Path", generated_fields.get("category_path_en")),
            ("Category Path ZH", generated_fields.get("category_path_zh")),
            ("Loadsheet Template ID", generated_fields.get("loadsheet_template_id")),
            ("Loadsheet Template Name", generated_fields.get("loadsheet_template_name")),
            ("Brand ID", generated_fields.get("brand_id")),
            ("Brand Name", generated_fields.get("brand_name")),
            ("SKU", generated_fields.get("sku")),
            ("Barcode", generated_fields.get("barcode")),
            ("Title", generated_fields.get("title")),
            ("Subtitle", generated_fields.get("subtitle")),
            ("Description", generated_fields.get("description")),
            ("What's in the Box", generated_fields.get("whats_in_the_box")),
            ("Selling Price", generated_fields.get("selling_price")),
            ("RRP", generated_fields.get("rrp")),
            ("Stock Quantity", generated_fields.get("stock_quantity")),
            ("Minimum Leadtime Days", generated_fields.get("minimum_leadtime_days")),
            ("Seller Warehouse ID", generated_fields.get("seller_warehouse_id")),
            ("Length CM", generated_fields.get("length_cm")),
            ("Width CM", generated_fields.get("width_cm")),
            ("Height CM", generated_fields.get("height_cm")),
            ("Weight G", generated_fields.get("weight_g")),
            ("Min Required Images", category.get("min_required_images")),
            ("Compliance Certificates", category.get("compliance_certificates") or []),
        ]
        self._append_key_value_sheet(product_sheet, rows)

        images_sheet = workbook.create_sheet("Images")
        images_sheet.append(["Position", "Image URL"])
        for index, image_url in enumerate(generated_fields.get("image_urls") or [], start=1):
            images_sheet.append([index, image_url])

        attributes_sheet = workbook.create_sheet("Dynamic Attributes")
        attributes_sheet.append(["Key", "Value", "Required", "Value Type"])
        for attribute in generated_fields.get("dynamic_attributes") or []:
            attributes_sheet.append(
                [
                    attribute.get("key"),
                    self._stringify(attribute.get("value")),
                    bool(attribute.get("required")),
                    attribute.get("value_type") or "",
                ]
            )

        validation_sheet = workbook.create_sheet("Validation")
        validation_sheet.append(["Level", "Field", "Message"])
        for issue in validation.get("issues") or []:
            validation_sheet.append([issue["level"], issue["field"], issue["message"]])
        if not validation.get("issues"):
            validation_sheet.append(["info", "loadsheet", "No validation issues found."])

        for sheet in workbook.worksheets:
            self._style_sheet(sheet)

        now = datetime.now(UTC)
        date_path = now.strftime("%Y/%m/%d")
        target_dir = self.storage_root / date_path
        target_dir.mkdir(parents=True, exist_ok=True)
        safe_sku = re.sub(r"[^A-Za-z0-9_.-]+", "-", request.sku.strip())[:48].strip("-") or "preview"
        file_name = f"takealot-loadsheet-{safe_sku}-{uuid4().hex[:12]}.xlsx"
        target_path = target_dir / file_name
        workbook.save(target_path)

        data = target_path.read_bytes()
        relative_path = self._relative_path(target_path)
        public_url = f"{self.public_base_url.rstrip('/')}/{relative_path}" if self.public_base_url else None
        return {
            "asset_id": None,
            "storage_path": str(target_path),
            "public_url": public_url,
            "content_type": LOADSHEET_CONTENT_TYPE,
            "size_bytes": len(data),
            "checksum_sha256": hashlib.sha256(data).hexdigest(),
        }

    def asset_payload_for_generated_loadsheet(self, loadsheet_asset: dict[str, Any]) -> dict[str, Any]:
        storage_path = loadsheet_asset.get("storage_path")
        file_name = Path(str(storage_path)).name if storage_path else None
        return {
            "asset_type": "loadsheet",
            "source": "generated",
            "original_file_name": file_name,
            "file_name": file_name,
            "storage_path": storage_path,
            "public_url": loadsheet_asset.get("public_url"),
            "external_url": None,
            "content_type": loadsheet_asset.get("content_type") or LOADSHEET_CONTENT_TYPE,
            "size_bytes": loadsheet_asset.get("size_bytes"),
            "checksum_sha256": loadsheet_asset.get("checksum_sha256"),
            "width": None,
            "height": None,
            "validation_status": "valid",
            "validation_errors": [],
            "raw_payload": {
                "generated_by": "listing_loadsheet_preview",
                "public_base_url_configured": bool(self.public_base_url),
            },
        }

    def submit_loadsheet_to_takealot(
        self,
        *,
        loadsheet_asset: dict[str, Any],
        api_key: str,
        submission_name: str | None = None,
    ) -> dict[str, Any]:
        if not api_key or not api_key.strip():
            raise ListingLoadsheetSubmitError("Store credentials unavailable")
        storage_path = Path(str(loadsheet_asset.get("storage_path") or ""))
        if not storage_path.exists() or not storage_path.is_file():
            raise ListingLoadsheetSubmitError("Generated loadsheet file is missing")
        file_name = submission_name or storage_path.name
        data = storage_path.read_bytes()
        try:
            with httpx.Client(timeout=self.submit_timeout_seconds) as client:
                response = client.post(
                    f"{self.submit_base_url}/loadsheets/submissions",
                    headers={
                        "Accept": "application/json",
                        "Authorization": f"Key {api_key.strip()}",
                        "Origin": "https://seller.takealot.com",
                        "Referer": "https://seller.takealot.com/",
                        "User-Agent": "Xiaohei-ERP/1.0",
                    },
                    files={
                        "loadsheet": (
                            file_name,
                            data,
                            loadsheet_asset.get("content_type") or LOADSHEET_CONTENT_TYPE,
                        )
                    },
                )
        except httpx.HTTPError as exc:
            safe_message = self.sanitize_text(str(exc), secrets=[api_key])
            raise ListingLoadsheetSubmitError(f"Takealot loadsheet submit request failed: {safe_message}") from exc

        payload = self.sanitize_official_response(self._response_payload(response), secrets=[api_key])
        if response.status_code in {401, 403}:
            raise ListingLoadsheetSubmitError(
                "Takealot rejected the store credentials.",
                status_code=response.status_code,
                official_response=payload,
            )
        if response.status_code in LOADSHEET_SUBMIT_RETRY_STATUS_CODES:
            raise ListingLoadsheetSubmitError(
                f"Takealot temporary loadsheet submit failure: HTTP {response.status_code}",
                status_code=response.status_code,
                official_response=payload,
            )
        if response.status_code >= 400:
            raise ListingLoadsheetSubmitError(
                self._extract_error_message(payload, fallback=f"Takealot loadsheet submit failed: HTTP {response.status_code}"),
                status_code=response.status_code,
                official_response=payload,
            )

        return {
            "takealot_submission_id": self.extract_submission_id(payload),
            "official_response": payload,
        }

    def list_submissions(
        self,
        *,
        api_key: str,
        limit: int = 50,
    ) -> dict[str, Any]:
        if not api_key or not api_key.strip():
            raise ListingLoadsheetStatusError("Store credentials unavailable")
        # This endpoint is used only from server-side workers. The API key is
        # placed in the request header and is never returned, logged, or stored.
        try:
            with httpx.Client(timeout=self.submit_timeout_seconds) as client:
                response = client.get(
                    f"{self.submit_base_url}/loadsheets/submissions",
                    params={"limit": max(1, min(int(limit), 100))},
                    headers=self._seller_loadsheet_headers(api_key),
                )
        except httpx.HTTPError as exc:
            safe_message = self.sanitize_text(str(exc), secrets=[api_key])
            raise ListingLoadsheetStatusError(f"Takealot loadsheet status request failed: {safe_message}") from exc
        return self._loadsheets_status_response_or_error(response, secrets=[api_key])

    def get_submission_status(
        self,
        *,
        api_key: str,
        takealot_submission_id: str,
    ) -> dict[str, Any]:
        if not api_key or not api_key.strip():
            raise ListingLoadsheetStatusError("Store credentials unavailable")
        submission_id = str(takealot_submission_id or "").strip()
        if not submission_id:
            raise ListingLoadsheetStatusError("Takealot submission id is required")
        # The status read is deliberately separate from offer finalization. A
        # loadsheet can be accepted for processing before content review is
        # actually approved, and creating an offer too early can produce
        # duplicate or rejected platform records.
        try:
            with httpx.Client(timeout=self.submit_timeout_seconds) as client:
                response = client.get(
                    f"{self.submit_base_url}/loadsheets/submissions/{quote(submission_id, safe='')}",
                    headers=self._seller_loadsheet_headers(api_key),
                )
        except httpx.HTTPError as exc:
            safe_message = self.sanitize_text(str(exc), secrets=[api_key])
            raise ListingLoadsheetStatusError(f"Takealot loadsheet status request failed: {safe_message}") from exc
        return self._loadsheets_status_response_or_error(response, secrets=[api_key])

    @classmethod
    def map_submission_status(cls, payload: dict[str, Any]) -> dict[str, Any]:
        statuses = cls._collect_status_values(payload)
        normalized_statuses = [cls._normalize_status_value(value) for value in statuses]
        official_status = next((value for value in statuses if value), "unknown")
        normalized_set = {value for value in normalized_statuses if value}

        # The mapping is conservative on purpose: only clear approval terms open
        # the Offer gate. Partial success is recorded for operators but does not
        # trigger offer creation because some lines may still be rejected.
        if normalized_set & TAKEALOT_APPROVED_REVIEW_STATUSES:
            return {
                "official_status": official_status,
                "status": "content_reviewed",
                "stage": "approved",
                "review_status": "approved",
                "approved": True,
                "terminal": True,
                "raw_statuses": statuses,
                "message": "Takealot loadsheet review is approved.",
            }
        if normalized_set & TAKEALOT_PARTIAL_REVIEW_STATUSES:
            return {
                "official_status": official_status,
                "status": "content_reviewed",
                "stage": "reviewed",
                "review_status": "partial",
                "approved": False,
                "terminal": True,
                "raw_statuses": statuses,
                "message": "Takealot loadsheet review is partially approved or needs changes.",
            }
        if normalized_set & TAKEALOT_REJECTED_REVIEW_STATUSES:
            local_status = "content_review_failed" if "failed" in normalized_set or "failure" in normalized_set else "content_rejected"
            return {
                "official_status": official_status,
                "status": local_status,
                "stage": "failed" if local_status == "content_review_failed" else "rejected",
                "review_status": "failed" if local_status == "content_review_failed" else "rejected",
                "approved": False,
                "terminal": True,
                "raw_statuses": statuses,
                "message": "Takealot loadsheet review did not pass.",
            }
        if normalized_set & TAKEALOT_PENDING_REVIEW_STATUSES:
            review_status = "under_review" if normalized_set & {"processing", "in_review", "under_review", "reviewing"} else "submitted"
            return {
                "official_status": official_status,
                "status": "content_submitted",
                "stage": "under_review" if review_status == "under_review" else "submitted",
                "review_status": review_status,
                "approved": False,
                "terminal": False,
                "raw_statuses": statuses,
                "message": "Takealot loadsheet review is still pending.",
            }
        return {
            "official_status": official_status,
            "status": "content_submitted",
            "stage": "submitted",
            "review_status": "unknown",
            "approved": False,
            "terminal": False,
            "raw_statuses": statuses,
            "message": "Takealot loadsheet review status is unknown.",
        }

    @classmethod
    def is_review_approved(cls, submission: dict[str, Any] | None) -> bool:
        return bool(submission and str(submission.get("review_status") or "").strip().lower() == "approved")

    @classmethod
    def sanitize_official_response(cls, value: Any, *, secrets: list[str] | None = None) -> Any:
        if isinstance(value, dict):
            sanitized: dict[str, Any] = {}
            for key, nested in value.items():
                key_text = str(key)
                lowered = key_text.lower()
                if any(part in lowered for part in SENSITIVE_RESPONSE_KEY_PARTS):
                    sanitized[key_text] = "[redacted]"
                else:
                    sanitized[key_text] = cls.sanitize_official_response(nested, secrets=secrets)
            return sanitized
        if isinstance(value, list):
            return [cls.sanitize_official_response(item, secrets=secrets) for item in value[:200]]
        if isinstance(value, str):
            return cls.sanitize_text(value, secrets=secrets)
        return value

    @classmethod
    def sanitize_text(cls, value: Any, *, secrets: list[str] | None = None) -> str:
        text = str(value or "")[:5000]
        for secret in secrets or []:
            secret_text = str(secret or "").strip()
            if len(secret_text) >= 8:
                text = text.replace(secret_text, "[redacted]")
        text = AUTHORIZATION_VALUE_RE.sub(r"\1: [redacted]", text)
        text = AUTHORIZATION_HEADER_RE.sub(r"\1: [redacted]", text)
        text = BEARER_TOKEN_RE.sub("Bearer [redacted]", text)
        text = KEY_TOKEN_RE.sub(lambda match: f"{match.group(1)} [redacted]", text)
        text = SECRET_ASSIGNMENT_RE.sub(lambda match: f"{match.group(1)}=[redacted]", text)
        text = SECRET_QUERY_RE.sub(lambda match: f"{match.group(1)}[redacted]", text)
        return LONG_SECRET_RE.sub("[redacted-secret]", text)

    def _collect_public_image_urls(
        self,
        *,
        request: ListingLoadsheetPreviewRequest,
        category: dict[str, Any],
        assets: list[dict[str, Any]],
        issues: list[dict[str, str]],
    ) -> dict[str, list[str]]:
        image_urls: list[str] = []
        for index, image_url in enumerate(request.image_urls):
            normalized_url = str(image_url or "").strip()
            if self._is_http_url(normalized_url):
                image_urls.append(normalized_url)
                continue
            issues.append(
                self._issue(
                    "error",
                    f"image_urls[{index}]",
                    "Loadsheet image URLs must be public http/https URLs.",
                )
            )

        assets_by_id = {str(asset.get("id")): asset for asset in assets if asset.get("id")}
        for asset_id in self._dedupe(request.asset_ids):
            asset = assets_by_id.get(asset_id)
            if asset is None:
                issues.append(self._issue("error", "asset_ids", f"Listing asset not found or not accessible: {asset_id}."))
                continue
            if str(asset.get("store_id") or "") != request.store_id:
                issues.append(self._issue("error", "asset_ids", f"Listing asset belongs to a different store: {asset_id}."))
                continue
            if asset.get("asset_type") != "image":
                issues.append(self._issue("error", "asset_ids", f"Listing asset is not an image: {asset_id}."))
                continue
            public_url = str(asset.get("public_url") or asset.get("external_url") or "").strip()
            if self._is_http_url(public_url):
                image_urls.append(public_url)
            else:
                issues.append(
                    self._issue(
                        "error",
                        "asset_ids",
                        f"Listing asset {asset_id} does not have a public http/https URL for Takealot loadsheets.",
                    )
                )

        image_urls = self._dedupe(image_urls)
        required_count = max(0, int(category.get("min_required_images") or 0))
        if len(image_urls) < required_count:
            issues.append(
                self._issue(
                    "error",
                    "image_urls",
                    f"{required_count} public image URL(s) required; {len(image_urls)} provided.",
                )
            )
        return {"image_urls": image_urls}

    def _validate_dynamic_attributes(
        self,
        *,
        raw_values: Any,
        allowed_attributes: list[dict[str, Any]],
        issues: list[dict[str, str]],
        missing_required_fields: list[str],
    ) -> list[dict[str, Any]]:
        allowed_by_key = {str(attribute.get("key")): attribute for attribute in allowed_attributes if attribute.get("key")}
        provided = self._dynamic_attributes_to_dict(raw_values)
        if provided is None:
            issues.append(
                self._issue(
                    "error",
                    "dynamic_attributes",
                    "dynamic_attributes must be an object or a list of key/value objects.",
                )
            )
            provided = {}

        normalized: list[dict[str, Any]] = []
        for key, value in provided.items():
            definition = allowed_by_key.get(key)
            if definition is None:
                issues.append(
                    self._issue(
                        "error",
                        f"dynamic_attributes.{key}",
                        "Dynamic attribute is not allowed for this category.",
                    )
                )
                continue
            coerced_value, valid = self._coerce_attribute_value(value, definition)
            if not valid:
                issues.append(
                    self._issue(
                        "error",
                        f"dynamic_attributes.{key}",
                        "Dynamic attribute value is not valid for the category definition.",
                    )
                )
                continue
            normalized.append(
                {
                    "key": key,
                    "value": coerced_value,
                    "required": bool(definition.get("required")),
                    "value_type": definition.get("value_type") or "text",
                }
            )

        for definition in allowed_attributes:
            if not definition.get("required"):
                continue
            key = str(definition.get("key") or "").strip()
            if not key:
                continue
            if key not in provided or self._is_blank(provided.get(key)):
                missing_required_fields.append(f"dynamic_attributes.{key}")
                issues.append(
                    self._issue(
                        "error",
                        f"dynamic_attributes.{key}",
                        "Required dynamic attribute is missing.",
                    )
                )
        return normalized

    @staticmethod
    def _append_key_value_sheet(sheet: Any, rows: list[tuple[str, Any]]) -> None:
        sheet.append(["Field", "Value"])
        for field, value in rows:
            sheet.append([field, ListingLoadsheetService._stringify(value)])

    @staticmethod
    def _style_sheet(sheet: Any) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            width = min(60, max(14, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            sheet.column_dimensions[column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _require_text(
        value: str,
        field: str,
        message: str,
        issues: list[dict[str, str]],
        missing_required_fields: list[str],
    ) -> None:
        if not str(value or "").strip():
            missing_required_fields.append(field)
            issues.append(ListingLoadsheetService._issue("error", field, message))

    @staticmethod
    def _positive_number(
        value: Any,
        field: str,
        issues: list[dict[str, str]],
        missing_required_fields: list[str],
    ) -> float | None:
        number = ListingLoadsheetService._number(value)
        if number is None:
            missing_required_fields.append(field)
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} is required."))
            return None
        if number <= 0:
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} must be greater than 0."))
            return None
        return number

    @staticmethod
    def _required_nonnegative_number(
        value: Any,
        field: str,
        issues: list[dict[str, str]],
        missing_required_fields: list[str],
    ) -> float | None:
        number = ListingLoadsheetService._number(value)
        if number is None:
            missing_required_fields.append(field)
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} is required."))
            return None
        if number < 0:
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} must be greater than or equal to 0."))
            return None
        return number

    @staticmethod
    def _required_nonnegative_int(
        value: Any,
        field: str,
        issues: list[dict[str, str]],
        missing_required_fields: list[str],
    ) -> int | None:
        number = ListingLoadsheetService._number(value)
        if number is None:
            missing_required_fields.append(field)
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} is required."))
            return None
        if number < 0:
            issues.append(ListingLoadsheetService._issue("error", field, f"{field} must be greater than or equal to 0."))
            return None
        return int(number)

    @staticmethod
    def _number(value: Any) -> float | None:
        if value is None or value == "" or isinstance(value, bool):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _dynamic_attributes_to_dict(raw_values: Any) -> dict[str, Any] | None:
        if raw_values in (None, ""):
            return {}
        if hasattr(raw_values, "model_dump"):
            raw_values = raw_values.model_dump()
        if isinstance(raw_values, dict):
            return {str(key): value for key, value in raw_values.items()}
        if isinstance(raw_values, list):
            result: dict[str, Any] = {}
            for item in raw_values:
                if hasattr(item, "model_dump"):
                    item = item.model_dump()
                if not isinstance(item, dict):
                    return None
                key = item.get("key") or item.get("name") or item.get("attribute_key")
                if key in (None, ""):
                    return None
                result[str(key)] = item.get("value")
            return result
        return None

    @staticmethod
    def _coerce_attribute_value(value: Any, definition: dict[str, Any]) -> tuple[Any, bool]:
        options = definition.get("options") or []
        value_type = definition.get("value_type") or "text"
        if options:
            for option in options:
                if str(value).strip().lower() == str(option).strip().lower():
                    return option, True
            return None, False
        if value_type == "boolean":
            if isinstance(value, bool):
                return "Yes" if value else "No", True
            normalized = str(value).strip().lower()
            if normalized in {"yes", "true", "1"}:
                return "Yes", True
            if normalized in {"no", "false", "0"}:
                return "No", True
            return None, False
        if value_type == "number":
            if value in (None, ""):
                return None, True
            try:
                return float(value), True
            except (TypeError, ValueError):
                return None, False
        return "" if value is None else str(value).strip(), True

    @staticmethod
    def _valid_whats_in_the_box(value: str) -> bool:
        lines = [line.strip() for line in value.replace("\r", "\n").split("\n") if line.strip()]
        return bool(lines) and all(re.match(r"^[1-9][0-9]*\s+x\s+.{2,}$", line) for line in lines)

    @staticmethod
    def _box_text(value: str) -> str:
        return "\n".join(
            re.sub(r"\s+", " ", line).strip()
            for line in value.replace("\r", "\n").split("\n")
            if line.strip()
        )

    @staticmethod
    def _is_http_url(value: str | None) -> bool:
        if not value:
            return False
        parsed = urlparse(value)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)

    @staticmethod
    def _is_blank(value: Any) -> bool:
        return value is None or str(value).strip() == ""

    @staticmethod
    def _dedupe(values: list[str]) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for value in values:
            normalized = str(value or "").strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            result.append(normalized)
        return result

    @staticmethod
    def _issue(level: str, field: str, message: str) -> dict[str, str]:
        return {"level": level, "field": field, "message": message}

    def _relative_path(self, target_path: Path) -> str:
        try:
            return target_path.relative_to(self.repo_root).as_posix()
        except ValueError:
            return target_path.name

    @staticmethod
    def _stringify(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    @staticmethod
    def _response_payload(response: httpx.Response) -> dict[str, Any]:
        try:
            payload = response.json()
        except ValueError:
            text = response.text[:500] if response.text else ""
            return {"message": text} if text else {}
        return payload if isinstance(payload, dict) else {"raw": payload}

    def _loadsheets_status_response_or_error(
        self,
        response: httpx.Response,
        *,
        secrets: list[str] | None = None,
    ) -> dict[str, Any]:
        payload = self.sanitize_official_response(self._response_payload(response), secrets=secrets)
        if response.status_code in {401, 403}:
            raise ListingLoadsheetStatusError(
                "Takealot rejected the store credentials.",
                status_code=response.status_code,
                official_response=payload,
            )
        if response.status_code in LOADSHEET_SUBMIT_RETRY_STATUS_CODES:
            raise ListingLoadsheetStatusError(
                f"Takealot temporary loadsheet status failure: HTTP {response.status_code}",
                status_code=response.status_code,
                official_response=payload,
            )
        if response.status_code >= 400:
            raise ListingLoadsheetStatusError(
                self._extract_error_message(payload, fallback=f"Takealot loadsheet status failed: HTTP {response.status_code}"),
                status_code=response.status_code,
                official_response=payload,
            )
        return payload

    @staticmethod
    def _seller_loadsheet_headers(api_key: str) -> dict[str, str]:
        return {
            "Accept": "application/json",
            "Authorization": f"Key {api_key.strip()}",
            "Origin": "https://seller.takealot.com",
            "Referer": "https://seller.takealot.com/",
            "User-Agent": "Xiaohei-ERP/1.0",
        }

    @classmethod
    def _collect_status_values(cls, payload: Any) -> list[str]:
        values: list[str] = []
        if not isinstance(payload, dict):
            return values
        direct_keys = (
            "submission_status",
            "submissionStatus",
            "review_status",
            "reviewStatus",
            "content_status",
            "contentStatus",
            "status",
            "state",
            "result",
        )
        for key in direct_keys:
            value = payload.get(key)
            if value not in (None, "") and not isinstance(value, (dict, list)):
                values.append(str(value))
        for key in ("data", "submission", "result", "loadsheet_response", "loadsheetResponse"):
            nested = payload.get(key)
            if isinstance(nested, dict):
                values.extend(cls._collect_status_values(nested))
        return values or ["unknown"]

    @staticmethod
    def _normalize_status_value(value: Any) -> str:
        text = str(value or "").strip().lower()
        return re.sub(r"[^a-z0-9]+", "_", text).strip("_")

    @staticmethod
    def _extract_error_message(payload: dict[str, Any], *, fallback: str) -> str:
        for key in ("message", "detail", "title", "error"):
            value = payload.get(key)
            if value:
                return str(value)[:500]
        return fallback

    @classmethod
    def extract_submission_id(cls, payload: Any) -> str:
        if not isinstance(payload, dict):
            return ""
        for key in (
            "submission_id",
            "submissionId",
            "id",
            "loadsheet_submission_id",
            "loadsheetSubmissionId",
            "takealot_submission_id",
            "takealotSubmissionId",
        ):
            value = payload.get(key)
            if value not in (None, ""):
                return str(value).strip()
        for key in ("data", "submission", "result", "loadsheet_response", "loadsheetResponse"):
            nested = cls.extract_submission_id(payload.get(key))
            if nested:
                return nested
        return ""
