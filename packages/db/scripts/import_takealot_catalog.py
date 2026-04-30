from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from psycopg.types.json import Jsonb

ROOT = Path(__file__).resolve().parents[3]
API_ROOT = ROOT / "apps" / "api"
sys.path.insert(0, str(API_ROOT))

from src.platform.db.session import get_db_session  # noqa: E402


CATEGORY_LABEL_RE = re.compile(r"^(?P<name>.*?)\s*\((?P<id>\d+)\)\s*$")
DEFAULT_CATEGORY_SHEET = "Loadsheet & Category Look-Up"
DEFAULT_IMAGE_REQUIREMENTS_SHEET = "Image Requirements"
DEFAULT_BRAND_SHEET = "Brand Look up"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", cell_text(value).lower())


def parse_category_label(value: Any) -> tuple[str, int]:
    text = cell_text(value)
    match = CATEGORY_LABEL_RE.match(text)
    if match:
        return match.group("name").strip(), int(match.group("id"))
    if text.isdigit():
        return text, int(text)
    return text, 0


def to_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def split_list(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [cell_text(item) for item in value if cell_text(item)]
    text = cell_text(value)
    if not text or text.lower() in {"n/a", "na", "none", "not applicable", "-"}:
        return []
    if text.startswith("["):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            return [cell_text(item) for item in parsed if cell_text(item)]
    return [part.strip() for part in re.split(r"[\n;,|]+", text) if part.strip()]


def build_path(*parts: str) -> str:
    return " > ".join(part for part in (cell_text(part) for part in parts) if part)


def build_search_text(*parts: Any) -> str:
    values = [cell_text(part) for part in parts if cell_text(part)]
    return " ".join(values)


def get_value(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        key = normalize_header(name)
        if key in row:
            return row[key]
    return None


def resolve_sheet(workbook: Any, requested: str, fallback_names: list[str]) -> str:
    if requested in workbook.sheetnames:
        return requested
    by_lower = {name.strip().lower(): name for name in workbook.sheetnames}
    for candidate in [requested, *fallback_names]:
        resolved = by_lower.get(candidate.strip().lower())
        if resolved:
            return resolved
    raise ValueError(f"Sheet not found: {requested}")


def iter_sheet_dicts(workbook: Any, sheet_name: str, header_row: int) -> Any:
    worksheet = workbook[sheet_name]
    header_values = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            values_only=True,
        )
    )
    headers = [normalize_header(value) for value in header_values]

    for row_values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {
            header: row_values[index] if index < len(row_values) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in row.values()):
            yield row


def category_key(
    division: str,
    department: str,
    main_category_id: int,
    category_id: int,
) -> tuple[str, str, int, int]:
    return (division, department, main_category_id, category_id)


def load_image_requirements(
    workbook: Any,
    *,
    sheet_name: str,
    header_row: int,
) -> dict[tuple[str, str, int, int], dict[str, Any]]:
    requirements: dict[tuple[str, str, int, int], dict[str, Any]] = {}
    if sheet_name not in workbook.sheetnames:
        return requirements

    for row in iter_sheet_dicts(workbook, sheet_name, header_row):
        division = cell_text(get_value(row, "Division"))
        department = cell_text(get_value(row, "Loadsheet/Department", "Department"))
        main_category_name, main_category_id = parse_category_label(get_value(row, "Main Category"))
        lowest_category_name, category_id = parse_category_label(get_value(row, "Lowest Category"))
        if not category_id:
            continue
        text_value = cell_text(
            get_value(
                row,
                "Images Requirements",
                "Image Requirements",
                "Minimum Required Images",
            )
        )
        min_images = to_int(text_value, 0)
        key = category_key(division, department, main_category_id, category_id)
        requirements[key] = {
            "division": division,
            "department": department,
            "main_category_id": main_category_id,
            "main_category_name": main_category_name,
            "category_id": category_id,
            "lowest_category_name": lowest_category_name,
            "lowest_category_raw": cell_text(get_value(row, "Lowest Category")),
            "min_required_images": min_images,
            "image_requirement_texts": [text_value] if text_value else [],
            "raw_payload": json_safe(row),
        }
    return requirements


def import_categories(
    connection: Any,
    workbook: Any,
    *,
    category_sheet: str,
    category_header_row: int,
    image_requirements_sheet: str,
    image_requirements_header_row: int,
    import_source: str,
) -> dict[str, int]:
    image_requirements = load_image_requirements(
        workbook,
        sheet_name=image_requirements_sheet,
        header_row=image_requirements_header_row,
    )
    created = 0
    updated = 0

    with connection.cursor() as cursor:
        existing_rows = cursor.execute(
            """
            select division, department, main_category_id, category_id
            from takealot_categories
            """
        ).fetchall()
        existing_keys = {
            category_key(
                row["division"],
                row["department"],
                int(row["main_category_id"]),
                int(row["category_id"]),
            )
            for row in existing_rows
        }

        touched_keys: set[tuple[str, str, int, int]] = set()
        for row in iter_sheet_dicts(workbook, category_sheet, category_header_row):
            division = cell_text(get_value(row, "Division"))
            department = cell_text(get_value(row, "Loadsheet/Department", "Department"))
            main_category_name, main_category_id = parse_category_label(get_value(row, "Main Category"))
            lowest_category_name, category_id = parse_category_label(get_value(row, "Lowest Category"))
            if not category_id:
                continue

            key = category_key(division, department, main_category_id, category_id)
            image_data = image_requirements.get(key, {})
            raw_lowest_category = cell_text(get_value(row, "Lowest Category"))
            min_required_images = int(
                image_data.get("min_required_images")
                if image_data.get("min_required_images") not in (None, 0)
                else to_int(get_value(row, "Minimum Required Images"), 1)
            )
            path_en = cell_text(get_value(row, "Path", "Path EN", "Category Path")) or build_path(
                division,
                department,
                main_category_name,
                lowest_category_name,
            )
            path_zh = cell_text(get_value(row, "Path ZH", "Chinese Path"))
            compliance_certificates = split_list(
                get_value(
                    row,
                    "Required Compliance Certificate",
                    "Compliance Certificate",
                    "Compliance Certificates",
                )
            )
            image_texts = image_data.get("image_requirement_texts") or split_list(
                get_value(row, "Image Requirement Text", "Image Requirements")
            )
            required_attributes = split_list(
                get_value(row, "Required Attributes", "Mandatory Attributes")
            )
            optional_attributes = split_list(get_value(row, "Optional Attributes"))
            template_id = cell_text(get_value(row, "Loadsheet Template ID", "Template ID"))
            template_name = cell_text(get_value(row, "Loadsheet Template Name", "Template Name"))
            search_text = build_search_text(
                category_id,
                division,
                department,
                main_category_id,
                main_category_name,
                lowest_category_name,
                raw_lowest_category,
                path_en,
                path_zh,
            )
            raw_payload = {
                "catalog_row": json_safe(row),
                "image_requirements_row": image_data.get("raw_payload"),
            }

            inserted = key not in existing_keys
            upsert_category(
                cursor,
                category_id=category_id,
                division=division,
                department=department,
                main_category_id=main_category_id,
                main_category_name=main_category_name,
                lowest_category_name=lowest_category_name,
                lowest_category_raw=raw_lowest_category,
                path_en=path_en,
                path_zh=path_zh,
                search_text=search_text,
                min_required_images=min_required_images,
                compliance_certificates=compliance_certificates,
                image_requirement_texts=image_texts,
                required_attributes=required_attributes,
                optional_attributes=optional_attributes,
                loadsheet_template_id=template_id or None,
                loadsheet_template_name=template_name or None,
                raw_payload=raw_payload,
                import_source=import_source,
            )
            if inserted:
                created += 1
                existing_keys.add(key)
            else:
                updated += 1
            touched_keys.add(key)

        for key, image_data in image_requirements.items():
            if key in touched_keys:
                continue
            division, department, main_category_id, category_id = key
            main_category_name = cell_text(image_data.get("main_category_name"))
            lowest_category_name = cell_text(image_data.get("lowest_category_name")) or str(category_id)
            lowest_category_raw = cell_text(image_data.get("lowest_category_raw")) or str(category_id)
            path_en = build_path(division, department, main_category_name, lowest_category_name)
            search_text = build_search_text(
                category_id,
                division,
                department,
                main_category_id,
                main_category_name,
                lowest_category_name,
                lowest_category_raw,
                path_en,
            )
            inserted = key not in existing_keys
            upsert_category(
                cursor,
                category_id=category_id,
                division=division,
                department=department,
                main_category_id=main_category_id,
                main_category_name=main_category_name,
                lowest_category_name=lowest_category_name,
                lowest_category_raw=lowest_category_raw,
                path_en=path_en,
                path_zh="",
                search_text=search_text,
                min_required_images=int(image_data.get("min_required_images") or 0),
                compliance_certificates=[],
                image_requirement_texts=image_data.get("image_requirement_texts") or [],
                required_attributes=[],
                optional_attributes=[],
                loadsheet_template_id=None,
                loadsheet_template_name=None,
                raw_payload={"image_requirements_row": image_data.get("raw_payload")},
                import_source=import_source,
            )
            if inserted:
                created += 1
                existing_keys.add(key)
            else:
                updated += 1

    return {"created": created, "updated": updated}


def upsert_category(cursor: Any, **payload: Any) -> None:
    cursor.execute(
        """
        insert into takealot_categories (
          category_id,
          division,
          department,
          main_category_id,
          main_category_name,
          lowest_category_name,
          lowest_category_raw,
          path_en,
          path_zh,
          search_text,
          min_required_images,
          compliance_certificates,
          image_requirement_texts,
          required_attributes,
          optional_attributes,
          loadsheet_template_id,
          loadsheet_template_name,
          raw_payload,
          import_source,
          imported_at
        )
        values (
          %(category_id)s,
          %(division)s,
          %(department)s,
          %(main_category_id)s,
          %(main_category_name)s,
          %(lowest_category_name)s,
          %(lowest_category_raw)s,
          %(path_en)s,
          %(path_zh)s,
          %(search_text)s,
          %(min_required_images)s,
          %(compliance_certificates)s,
          %(image_requirement_texts)s,
          %(required_attributes)s,
          %(optional_attributes)s,
          %(loadsheet_template_id)s,
          %(loadsheet_template_name)s,
          %(raw_payload)s,
          %(import_source)s,
          now()
        )
        on conflict (division, department, main_category_id, category_id)
        do update set
          main_category_name = excluded.main_category_name,
          lowest_category_name = excluded.lowest_category_name,
          lowest_category_raw = excluded.lowest_category_raw,
          path_en = excluded.path_en,
          path_zh = excluded.path_zh,
          search_text = excluded.search_text,
          min_required_images = excluded.min_required_images,
          compliance_certificates = excluded.compliance_certificates,
          image_requirement_texts = excluded.image_requirement_texts,
          required_attributes = excluded.required_attributes,
          optional_attributes = excluded.optional_attributes,
          loadsheet_template_id = excluded.loadsheet_template_id,
          loadsheet_template_name = excluded.loadsheet_template_name,
          raw_payload = excluded.raw_payload,
          import_source = excluded.import_source,
          imported_at = now(),
          updated_at = now()
        """,
        {
            **payload,
            "compliance_certificates": Jsonb(payload["compliance_certificates"]),
            "image_requirement_texts": Jsonb(payload["image_requirement_texts"]),
            "required_attributes": Jsonb(payload["required_attributes"]),
            "optional_attributes": Jsonb(payload["optional_attributes"]),
            "raw_payload": Jsonb(json_safe(payload["raw_payload"])),
        },
    )


def import_brands(
    connection: Any,
    workbook: Any,
    *,
    brand_sheet: str,
    brand_header_row: int,
    import_source: str,
) -> dict[str, int]:
    created = 0
    updated = 0

    with connection.cursor() as cursor:
        existing_rows = cursor.execute("select brand_id from takealot_brands").fetchall()
        existing_ids = {row["brand_id"] for row in existing_rows}

        for row in iter_sheet_dicts(workbook, brand_sheet, brand_header_row):
            brand_id = cell_text(get_value(row, "brand_id", "Brand ID", "ID"))
            brand_name = cell_text(get_value(row, "brand_name", "Brand Name", "Name"))
            if not brand_id or not brand_name:
                continue
            inserted = brand_id not in existing_ids
            cursor.execute(
                """
                insert into takealot_brands (
                  brand_id,
                  brand_name,
                  search_text,
                  raw_payload,
                  import_source,
                  imported_at
                )
                values (%s, %s, %s, %s, %s, now())
                on conflict (brand_id)
                do update set
                  brand_name = excluded.brand_name,
                  search_text = excluded.search_text,
                  raw_payload = excluded.raw_payload,
                  import_source = excluded.import_source,
                  imported_at = now(),
                  updated_at = now()
                """,
                (
                    brand_id,
                    brand_name,
                    build_search_text(brand_id, brand_name),
                    Jsonb({"brand_row": json_safe(row)}),
                    import_source,
                ),
            )
            if inserted:
                created += 1
                existing_ids.add(brand_id)
            else:
                updated += 1

    return {"created": created, "updated": updated}


def import_workbook(args: argparse.Namespace) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required. Install the API package dependencies first.") from exc

    workbook_path = Path(args.file).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        with get_db_session() as connection:
            summary = {
                "file": str(workbook_path),
                "dry_run": args.dry_run,
                "categories": {"created": 0, "updated": 0},
                "brands": {"created": 0, "updated": 0},
            }
            if args.only in {"all", "categories"}:
                category_sheet = resolve_sheet(
                    workbook,
                    args.category_sheet,
                    [DEFAULT_CATEGORY_SHEET],
                )
                try:
                    image_requirements_sheet = resolve_sheet(
                        workbook,
                        args.image_requirements_sheet,
                        [DEFAULT_IMAGE_REQUIREMENTS_SHEET],
                    )
                except ValueError:
                    image_requirements_sheet = args.image_requirements_sheet
                summary["categories"] = import_categories(
                    connection,
                    workbook,
                    category_sheet=category_sheet,
                    category_header_row=args.category_header_row,
                    image_requirements_sheet=image_requirements_sheet,
                    image_requirements_header_row=args.image_requirements_header_row,
                    import_source=args.import_source,
                )
            if args.only in {"all", "brands"}:
                brand_sheet = resolve_sheet(workbook, args.brand_sheet, [DEFAULT_BRAND_SHEET])
                summary["brands"] = import_brands(
                    connection,
                    workbook,
                    brand_sheet=brand_sheet,
                    brand_header_row=args.brand_header_row,
                    import_source=args.import_source,
                )
            if args.dry_run:
                connection.rollback()
            else:
                connection.commit()
            return summary
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import Takealot category and brand catalog Excel data into PostgreSQL."
    )
    parser.add_argument("--file", required=True, help="Path to the Takealot catalog xlsx file.")
    parser.add_argument("--only", choices=["all", "categories", "brands"], default="all")
    parser.add_argument("--category-sheet", default=DEFAULT_CATEGORY_SHEET)
    parser.add_argument("--category-header-row", type=int, default=3)
    parser.add_argument("--image-requirements-sheet", default=DEFAULT_IMAGE_REQUIREMENTS_SHEET)
    parser.add_argument("--image-requirements-header-row", type=int, default=3)
    parser.add_argument("--brand-sheet", default=DEFAULT_BRAND_SHEET)
    parser.add_argument("--brand-header-row", type=int, default=1)
    parser.add_argument("--import-source", default="takealot_catalog_excel")
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> None:
    summary = import_workbook(parse_args())
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
