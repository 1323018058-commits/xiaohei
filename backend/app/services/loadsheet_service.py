"""Takealot loadsheet Excel filling service.

Takes a downloaded .xlsm template + AI-rewritten data + job info,
fills in the loadsheet and returns the completed Excel as bytes.
"""
from __future__ import annotations

import io
import logging
import uuid
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

DEFAULT_WEIGHT_KG = 0.5
DEFAULT_DIMENSIONS_CM = [20.0, 15.0, 10.0]


def _build_internal_sku(job: dict[str, Any]) -> str:
    """Generate a unique SKU for the loadsheet."""
    asin = job.get("asin", "")
    jid = job.get("id", "")
    if asin:
        return f"DS-{asin}-{jid}"
    return f"DS-{uuid.uuid4().hex[:8].upper()}-{jid}"


def _normalize_barcode(raw: Any) -> str:
    """Extract a valid barcode (digits only, 8-14 chars)."""
    if not raw:
        return ""
    s = str(raw).strip().replace("-", "").replace(" ", "")
    # Only return if it looks like a real barcode
    if s.isdigit() and 8 <= len(s) <= 14:
        return s
    return ""


def fill_loadsheet_excel(
    template_bytes: bytes,
    ai: dict[str, Any],
    job: dict[str, Any],
) -> bytes:
    """Fill a downloaded Takealot loadsheet Excel with AI-generated data.

    Args:
        template_bytes: Raw .xlsm file bytes from Takealot API.
        ai: Dict from ai_analyze_and_rewrite() with listing fields.
        job: Dict with job-level data (asin, price_zar, weight_kg, barcode, image_url, etc.)

    Returns:
        Filled Excel file as bytes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    ws = wb["Loadsheet"]

    # Read column map from row 1 (API field names)
    col_map: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column
    # Also check row 4 for alternate key names
    for cell in ws[4]:
        if cell.value and str(cell.value).strip() not in col_map:
            col_map[str(cell.value).strip()] = cell.column

    # Find first empty data row starting from row 7
    data_row = 7
    for row in range(7, 20):
        sku_cell = ws.cell(row=row, column=col_map.get("SKU", 2)).value
        if not sku_cell:
            data_row = row
            break

    def set_cell(field: str, value: Any) -> None:
        col = col_map.get(field)
        if col and value is not None and value != "":
            ws.cell(row=data_row, column=col, value=value)

    def set_any(fields: list[str], value: Any) -> None:
        for field in fields:
            col = col_map.get(field)
            if col and value is not None and value != "":
                ws.cell(row=data_row, column=col, value=value)
                return

    sku = job.get("sku") or _build_internal_sku(job)
    barcode = _normalize_barcode(job.get("barcode"))

    # ── Core fields ──
    set_cell("Variant.ProductVariant", "Product")
    set_cell("SKU", sku)
    set_cell("TopCategory", ai.get("top_category", ""))
    set_cell("Category", ai.get("lowest_category", ""))

    if barcode:
        if "ProductID" in col_map:
            set_cell("ProductID", barcode)
        else:
            set_cell("ProductID.Value", barcode)

    set_cell("title", ai.get("listing_title", "")[:75])
    set_cell("subtitle", ai.get("subtitle", "")[:60])
    set_cell("description", ai.get("listing_description", ""))
    set_cell("Attribute.whats_in_the_box", ai.get("package_contents", ""))

    brand = ai.get("brand", "") or ""
    if brand.lower() == "generic":
        brand = ""
    if brand:
        set_cell("Brand", brand)

    set_cell("Attribute.model_number", ai.get("model_number", ""))
    set_cell("color.main", ai.get("color_main", ""))

    # ── Pricing ──
    price = int(float(job.get("price_zar") or 0))
    if price > 0:
        set_cell("SuggestedPrice.Amount", price)
        set_cell("SuggestedPrice.Currency", "ZAR")

    # ── Dimensions / Weight ──
    wg = ai.get("weight_grams")
    pwg = ai.get("packaged_weight_grams")
    weight_kg = float(job.get("weight_kg") or DEFAULT_WEIGHT_KG)
    dims = job.get("package_dimensions_cm") or DEFAULT_DIMENSIONS_CM

    if isinstance(dims, str):
        try:
            dims = [float(v.strip()) for v in dims.split(",")]
        except Exception:
            dims = DEFAULT_DIMENSIONS_CM

    try:
        dim_l, dim_w, dim_h = [round(float(v), 2) for v in dims]
    except Exception:
        dim_l, dim_w, dim_h = DEFAULT_DIMENSIONS_CM

    if not wg:
        wg = int(weight_kg * 1000)
    if not pwg:
        pwg = int(wg * 1.2)

    set_cell("Attribute.merchant_packaged_weight.value", pwg)
    set_cell("Attribute.merchant_packaged_weight.unit", "g")

    set_any([
        "Attribute.merchant_packaged_dimensions.length.value",
        "Attribute.merchant_packaged_dimensions.length",
        "Attribute.merchant_packaged_length.value",
        "Attribute.package_dimensions.length.value",
    ], dim_l)
    set_any([
        "Attribute.merchant_packaged_dimensions.width.value",
        "Attribute.merchant_packaged_dimensions.width",
        "Attribute.merchant_packaged_width.value",
        "Attribute.package_dimensions.width.value",
    ], dim_w)
    set_any([
        "Attribute.merchant_packaged_dimensions.height.value",
        "Attribute.merchant_packaged_dimensions.height",
        "Attribute.merchant_packaged_height.value",
        "Attribute.package_dimensions.height.value",
    ], dim_h)

    for dim_type in ["length", "width", "height"]:
        set_any([
            f"Attribute.merchant_packaged_dimensions.{dim_type}.unit",
            f"Attribute.merchant_packaged_{dim_type}.unit",
            f"Attribute.package_dimensions.{dim_type}.unit",
        ], "cm")

    # ── Warranty ──
    wm = ai.get("warranty_months")
    if wm:
        set_cell("Attribute.warranty.period.value", wm)
        set_cell("Attribute.warranty.period.unit", "m")

    # ── Boolean attributes ──
    bool_fields = [
        "Attribute.is_wireless", "Attribute.is_portable", "Attribute.is_waterproof",
        "Attribute.is_water_resistant", "Attribute.is_rechargeable",
        "Attribute.has_bluetooth", "Attribute.has_noise_cancelling",
        "Attribute.has_gps", "Attribute.is_lightweight", "Attribute.is_ergonomic",
        "Attribute.has_memory_card_slot", "Attribute.fast_charging",
        "Attribute.is_foldable", "Attribute.is_adjustable",
    ]
    for f in bool_fields:
        short = f.replace("Attribute.", "")
        val = ai.get(short)
        if val is not None:
            set_cell(f, val)

    # ── Image URL ──
    image_url = job.get("image_url", "")
    if image_url:
        for field in ["Image.1", "$Images", "images.[0]"]:
            set_cell(field, image_url)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
