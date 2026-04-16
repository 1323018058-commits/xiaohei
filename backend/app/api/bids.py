"""Bid management API router — engine control, products, log, insights."""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, case

from app.api.deps import ActiveUser, DbSession, RedisConn
from app.models.product import BidProduct
from app.schemas.bid import BidProductPatch, BidProductUpsert
from app.schemas.common import OkResponse
from app.services import bid_service, product_sync_progress_service, store_service

router = APIRouter(prefix="/api/bids/{store_id}", tags=["bids"])
logger = logging.getLogger(__name__)

_SYNC_STATUS_CACHE_TTL_SECONDS = 3
_BID_STATUS_CACHE_TTL_SECONDS = 10
_CHINA_TZ = timezone(timedelta(hours=8))


def _sync_status_cache_key(store_id: int) -> str:
    return f"cache:bid_product_sync_status:{store_id}"


def _bid_status_cache_key(store_id: int) -> str:
    return f"cache:bid_status:{store_id}"


def _format_china_time(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            try:
                dt = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return text
    else:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_CHINA_TZ).strftime("%Y-%m-%d %H:%M:%S")


async def _get_cached_json(redis, key: str) -> dict | None:
    raw = await redis.get(key)
    if raw is None:
        return None
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError, UnicodeDecodeError):
        await redis.delete(key)
        return None


async def _set_cached_json(redis, key: str, payload: dict, ttl_seconds: int) -> None:
    await redis.setex(key, ttl_seconds, json.dumps(payload, default=str))


async def _delete_cache_key(redis, key: str) -> None:
    try:
        await redis.delete(key)
    except Exception:
        logger.warning("failed to delete cache key %s", key, exc_info=True)


async def _require_store(db, store_id: int, user_id: int):
    store = await store_service.get_store(db, store_id, user_id)
    if not store:
        raise HTTPException(status_code=404, detail="店铺不存在")
    return store


# ---------------------------------------------------------------------------
# Engine control
# ---------------------------------------------------------------------------

@router.get("/status")
async def bid_status(store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn):
    store = await _require_store(db, store_id, user.id)
    cache_key = _bid_status_cache_key(store_id)
    cached = await _get_cached_json(redis, cache_key)
    if cached is not None:
        return cached

    state = await bid_service.get_engine_state(db, store_id)

    # Single query for both total and active counts
    result = await db.execute(
        select(
            func.count().label("total"),
            func.sum(case((BidProduct.auto_bid_enabled == 1, 1), else_=0)).label("active"),
        ).where(BidProduct.store_binding_id == store_id)
    )
    row = result.one()
    total = row.total or 0
    active = row.active or 0
    state["total_products"] = total
    state["active_products"] = active
    state["paused_products"] = total - active
    state["last_run"] = _format_china_time(state.get("last_run"))
    state["next_run"] = _format_china_time(state.get("next_run"))
    state["last_product_sync_at"] = _format_china_time(
        getattr(store, "last_synced_at", None)
    )
    state["next_product_sync_at"] = _format_china_time(
        store.last_synced_at + timedelta(minutes=30)
        if getattr(store, "last_synced_at", None)
        else None
    )

    payload = {"ok": True, "state": state}
    await _set_cached_json(redis, cache_key, payload, _BID_STATUS_CACHE_TTL_SECONDS)
    return payload


@router.post("/start", response_model=OkResponse)
async def bid_start(store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn):
    store = await _require_store(db, store_id, user.id)
    state = await bid_service.get_engine_state(db, store_id)
    if state["running"]:
        raise HTTPException(status_code=409, detail="出价引擎已在运行")
    await bid_service.set_engine_running(db, store_id, True)
    await db.commit()
    await _delete_cache_key(redis, _bid_status_cache_key(store_id))
    return OkResponse()


@router.post("/stop", response_model=OkResponse)
async def bid_stop(store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn):
    await _require_store(db, store_id, user.id)
    await bid_service.set_engine_running(db, store_id, False)
    await db.commit()
    await _delete_cache_key(redis, _bid_status_cache_key(store_id))
    return OkResponse()


# ---------------------------------------------------------------------------
# Products
# ---------------------------------------------------------------------------

@router.get("/products")
async def bid_products(
    store_id: int, user: ActiveUser, db: DbSession,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    sku: str = "",
    enabled: str = "",
    status: str = "",
):
    store = await _require_store(db, store_id, user.id)
    products, total = await bid_service.list_bid_products(
        db, store_id, page=page, page_size=page_size, sku=sku, enabled=enabled, status=status,
    )
    items = []
    for p in products:
        items.append({
            "id": p.id, "offer_id": p.offer_id, "sku": p.sku, "plid": p.plid,
            "title": p.title, "floor_price_zar": p.floor_price_zar,
            "target_price_zar": p.target_price_zar, "current_price_zar": p.current_price_zar,
            "buybox_price_zar": bid_service.resolve_buybox_display_price(
                p.current_price_zar,
                p.buybox_price_zar,
                store=store,
                buybox_store=p.buybox_store,
            ),
            "auto_bid_enabled": p.auto_bid_enabled,
            "last_action": p.last_action, "brand": p.brand, "image_url": p.image_url,
            "takealot_url": bid_service.resolve_takealot_url(p.takealot_url, p.plid),
            "api_status": p.api_status, "offer_status": p.offer_status,
            "last_checked_at": str(p.last_checked_at) if p.last_checked_at else None,
        })
    return {"ok": True, "total": total, "page": page, "page_size": page_size, "products": items}


@router.post("/products/{offer_id}/refresh-buybox")
async def bid_refresh_buybox(
    store_id: int,
    offer_id: str,
    user: ActiveUser,
    db: DbSession,
):
    store = await _require_store(db, store_id, user.id)
    product = await bid_service.get_bid_product(db, store_id, offer_id)
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    if not product.plid:
        raise HTTPException(status_code=409, detail="该商品缺少 PLID，无法刷新 BuyBox")

    from app.services.buybox_service import fetch_product_detail

    detail = await fetch_product_detail(product.plid)
    if not detail.get("ok"):
        raise HTTPException(status_code=502, detail=detail.get("error") or "BuyBox 刷新失败")
    if detail.get("buybox_price") is None:
        raise HTTPException(status_code=502, detail="未获取到 BuyBox 价格")

    refreshed_buybox = bid_service.resolve_buybox_display_price(
        product.current_price_zar,
        detail.get("buybox_price"),
        store=store,
        buybox_store=detail.get("buybox_seller") or product.buybox_store,
    )
    if refreshed_buybox is not None:
        product.buybox_price_zar = refreshed_buybox
    if detail.get("buybox_seller"):
        product.buybox_store = str(detail.get("buybox_seller"))

    resolved_url = bid_service.resolve_takealot_url(detail.get("takealot_url"), product.plid)
    if resolved_url:
        product.takealot_url = resolved_url

    product.last_checked_at = datetime.utcnow()
    await db.flush()

    return {
        "ok": True,
        "product": {
            "id": product.id,
            "offer_id": product.offer_id,
            "sku": product.sku,
            "plid": product.plid,
            "title": product.title,
            "current_price_zar": product.current_price_zar,
            "buybox_price_zar": product.buybox_price_zar,
            "buybox_store": product.buybox_store,
            "takealot_url": bid_service.resolve_takealot_url(product.takealot_url, product.plid),
            "last_checked_at": str(product.last_checked_at) if product.last_checked_at else None,
            "next_offer_price": detail.get("next_offer_price"),
        },
    }


@router.post("/products/refresh-buybox-all")
async def bid_refresh_buybox_all(
    store_id: int,
    user: ActiveUser,
    db: DbSession,
):
    store = await _require_store(db, store_id, user.id)
    result = await bid_service.refresh_store_buybox(db, store)
    return {"ok": True, **result}


@router.post("/products/sync")
async def bid_sync(store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn):
    store = await _require_store(db, store_id, user.id)
    def _load_task():
        from app.tasks.product_sync_tasks import run_bid_product_sync
        return run_bid_product_sync

    try:
        response = await product_sync_progress_service.enqueue_sync(
            redis,
            store_id,
            task_importer=_load_task,
            progress_scope="bids",
            lock_scope=product_sync_progress_service.DEFAULT_LOCK_SCOPE,
            queued_message="自动出价商品同步任务已提交，等待开始...",
            conflict_message="商品管理同步正在运行，请稍后再试",
        )
        await _delete_cache_key(redis, _sync_status_cache_key(store_id))
        return response
    except product_sync_progress_service.SyncTaskMissing:
        raise HTTPException(status_code=503, detail="同步任务未就绪")


@router.get("/products/sync/status")
async def bid_sync_status(store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn):
    await _require_store(db, store_id, user.id)
    cache_key = _sync_status_cache_key(store_id)
    cached = await _get_cached_json(redis, cache_key)
    if cached is not None:
        return cached

    progress = await product_sync_progress_service.get_progress(
        redis,
        store_id,
        scope="bids",
    )
    payload = {"ok": True, **progress}
    await _set_cached_json(redis, cache_key, payload, _SYNC_STATUS_CACHE_TTL_SECONDS)
    return payload


@router.post("/products")
async def bid_upsert_product(
    store_id: int,
    user: ActiveUser,
    db: DbSession,
    redis: RedisConn,
    body: BidProductUpsert,
):
    await _require_store(db, store_id, user.id)
    payload = body.model_dump()
    offer_id = payload.get("offer_id", "")
    if not offer_id:
        raise HTTPException(status_code=400, detail="缺少 offer_id")
    await bid_service.upsert_bid_product(db, store_id, payload)
    await db.commit()
    await _delete_cache_key(redis, _bid_status_cache_key(store_id))
    return OkResponse()


@router.patch("/products/{offer_id}")
async def bid_patch_product(
    store_id: int,
    offer_id: str,
    user: ActiveUser,
    db: DbSession,
    redis: RedisConn,
    body: BidProductPatch,
):
    await _require_store(db, store_id, user.id)
    payload = body.model_dump(exclude_unset=True)
    product = await bid_service.patch_bid_product(db, store_id, offer_id, **payload)
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    await db.commit()
    await _delete_cache_key(redis, _bid_status_cache_key(store_id))
    return OkResponse()


# ---------------------------------------------------------------------------
# Log & Insights
# ---------------------------------------------------------------------------

@router.get("/log")
async def bid_log(
    store_id: int, user: ActiveUser, db: DbSession,
    limit: int = Query(100, ge=1, le=500),
):
    await _require_store(db, store_id, user.id)
    logs = await bid_service.list_bid_log(db, store_id, limit=limit)
    items = []
    for log in logs:
        items.append({
            "id": log.id, "offer_id": log.offer_id, "sku": log.sku,
            "old_price": log.old_price, "new_price": log.new_price,
            "buybox_price": log.buybox_price, "action": log.action,
            "reason": log.reason,
            "created_at": str(log.created_at) if log.created_at else None,
        })
    return {"ok": True, "log": items}


@router.get("/insights")
async def bid_insights(store_id: int, user: ActiveUser, db: DbSession):
    await _require_store(db, store_id, user.id)
    insights = await bid_service.get_bid_insights(db, store_id)
    return {"ok": True, **insights}


# ---------------------------------------------------------------------------
# CSV/XLSX Export & Import
# ---------------------------------------------------------------------------

@router.get("/products/export")
async def bid_export(store_id: int, user: ActiveUser, db: DbSession):
    """Export bid products as CSV file."""
    await _require_store(db, store_id, user.id)
    products, _ = await bid_service.list_bid_products(db, store_id, page=1, page_size=10000)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "offer_id", "sku", "plid", "title", "brand",
        "floor_price_zar", "target_price_zar", "current_price_zar",
        "buybox_price_zar", "auto_bid_enabled",
    ])
    for p in products:
        writer.writerow([
            p.offer_id, p.sku, p.plid, p.title, p.brand,
            p.floor_price_zar, p.target_price_zar, p.current_price_zar,
            p.buybox_price_zar, p.auto_bid_enabled,
        ])

    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=bid_products_store{store_id}.csv"},
    )


@router.post("/products/import")
async def bid_import(
    store_id: int, user: ActiveUser, db: DbSession, redis: RedisConn, file: UploadFile = File(...),
):
    """Import bid products from CSV/XLSX file."""
    await _require_store(db, store_id, user.id)

    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = dict(zip(headers, row))
                if d.get("offer_id"):
                    rows.append(d)
            wb.close()
        except ImportError:
            raise HTTPException(status_code=400, detail="服务器未安装 openpyxl，请上传 CSV 格式")
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for d in reader:
            if d.get("offer_id"):
                rows.append(d)

    if not rows:
        raise HTTPException(status_code=400, detail="文件中无有效数据")

    imported = 0
    for d in rows[:5000]:
        offer_id = str(d.get("offer_id", "")).strip()
        if not offer_id:
            continue
        product_data = {
            "offer_id": offer_id,
            "sku": str(d.get("sku", "")).strip(),
            "plid": str(d.get("plid", "")).strip(),
            "title": str(d.get("title", "")).strip(),
            "brand": str(d.get("brand", "")).strip(),
        }
        floor = d.get("floor_price_zar")
        target = d.get("target_price_zar")
        enabled = d.get("auto_bid_enabled")
        if floor is not None and str(floor).strip():
            try:
                product_data["floor_price_zar"] = float(floor)
            except (ValueError, TypeError):
                pass
        if target is not None and str(target).strip():
            try:
                product_data["target_price_zar"] = float(target)
            except (ValueError, TypeError):
                pass
        if enabled is not None and str(enabled).strip():
            product_data["auto_bid_enabled"] = 1 if str(enabled).strip() in ("1", "true", "True", "是") else 0

        await bid_service.upsert_bid_product(db, store_id, product_data)
        imported += 1

    await db.commit()
    await _delete_cache_key(redis, _bid_status_cache_key(store_id))
    return {"ok": True, "imported": imported, "total_rows": len(rows)}
